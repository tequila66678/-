from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from ..database import get_db
from ..models import Score, Student, SportEvent, ScoringStandard, Class, Admin, SystemConfig, InputFormat, Gender
from ..schemas import ScoreBatchSave, ScoreWithChange, ClearAllRequest
from ..auth import get_current_admin, get_current_admin_flexible, get_super_admin, verify_password
from ..scoring import calculate_score, normalize_time_ms
import openpyxl
from io import BytesIO
from datetime import date
from typing import Optional
from collections import defaultdict

router = APIRouter(prefix="/api/scores", tags=["scores"])

def _get_previous_score(db: Session, student_db_id: int, event_id: int, current_date: date) -> Optional[Score]:
    return (
        db.query(Score)
        .filter(
            Score.student_id == student_db_id,
            Score.event_id == event_id,
            Score.test_date < current_date
        )
        .order_by(Score.test_date.desc())
        .first()
    )

@router.post("/batch", response_model=list[ScoreWithChange])
def batch_save_scores(
    data: ScoreBatchSave,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    results = []
    praise_threshold = 1
    warning_threshold = 2
    praise_cfg = db.query(SystemConfig).filter(SystemConfig.key == "praise_threshold").first()
    warning_cfg = db.query(SystemConfig).filter(SystemConfig.key == "warning_threshold").first()
    if praise_cfg:
        praise_threshold = int(praise_cfg.value)
    if warning_cfg:
        warning_threshold = int(warning_cfg.value)

    for entry in data.scores:
        event = db.query(SportEvent).get(entry.event_id)
        student = db.query(Student).get(entry.student_id)
        raw_value = entry.raw_value
        if event and event.input_format == InputFormat.time_ms:
            raw_value = normalize_time_ms(raw_value)
        standards = db.query(ScoringStandard).filter(ScoringStandard.event_id == entry.event_id).all()
        earned = calculate_score(raw_value, event, standards, student.gender.value if student else None)

        prev = _get_previous_score(db, entry.student_id, entry.event_id, entry.test_date)
        prev_score = None
        change = None
        is_praise = False
        is_warning = False
        if prev:
            prev_score = prev.earned_score
            change = earned - prev_score
            is_praise = change >= praise_threshold
            is_warning = (prev_score - earned) >= warning_threshold

        existing = (
            db.query(Score)
            .filter(
                Score.student_id == entry.student_id,
                Score.event_id == entry.event_id,
                Score.test_date == entry.test_date
            ).first()
        )
        if existing:
            existing.raw_value = raw_value
            existing.earned_score = earned
            existing.recorder_id = current.id
            score_obj = existing
        else:
            score_obj = Score(
                student_id=entry.student_id,
                event_id=entry.event_id,
                raw_value=raw_value,
                earned_score=earned,
                test_date=entry.test_date,
                recorder_id=current.id
            )
            db.add(score_obj)

        db.flush()
        db.refresh(score_obj)

        result = ScoreWithChange(
            id=score_obj.id,
            student_id=score_obj.student_id,
            event_id=score_obj.event_id,
            raw_value=score_obj.raw_value,
            earned_score=score_obj.earned_score,
            test_date=score_obj.test_date,
            previous_score=prev_score,
            change=change,
            is_praise=is_praise,
            is_warning=is_warning,
        )
        results.append(result)

    db.commit()
    return results

@router.delete("/{score_id}")
def delete_score(score_id: int, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sc = db.query(Score).get(score_id)
    if not sc:
        raise HTTPException(404, "成绩记录不存在")
    db.delete(sc)
    db.commit()
    return {"ok": True}

@router.get("/class-stats")
def class_stats(
    class_id: int = Query(...),
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    cls = db.query(Class).get(class_id)
    if not cls:
        raise HTTPException(404, "班级不存在")

    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    students = db.query(Student).filter(Student.class_id == class_id).all()
    scores_q = db.query(Score).filter(
        Score.student_id.in_([s.id for s in students])
    )
    if event_id_list:
        scores_q = scores_q.filter(Score.event_id.in_(event_id_list))

    all_scores = scores_q.order_by(Score.test_date.desc()).all()
    latest = {}
    for sc in all_scores:
        key = (sc.student_id, sc.event_id)
        if key not in latest:
            latest[key] = sc

    events = db.query(SportEvent).all()
    if event_id_list:
        events = [e for e in events if e.id in event_id_list]

    event_scores = defaultdict(list)
    student_totals = defaultdict(list)
    for (sid, eid), sc in latest.items():
        event_scores[eid].append(sc.earned_score)
        student_totals[sid].append(sc.earned_score)

    event_avgs = []
    for e in events:
        scores_list = event_scores.get(e.id, [])
        avg = sum(scores_list) / len(scores_list) if scores_list else 0
        event_avgs.append({"event_id": e.id, "event_name": e.name, "avg_score": round(avg, 1)})

    total_scores = [sum(v) for v in student_totals.values() if v]
    max_per_student = len(events)
    overall_avg = sum(total_scores) / len(total_scores) if total_scores else 0
    excellent_count = sum(1 for t in total_scores if max_per_student > 0 and t / max_per_student >= 9)
    pass_count = sum(1 for t in total_scores if max_per_student > 0 and t / max_per_student >= 6)
    n_students = len(student_totals)

    warning_students = []
    for s in students:
        for e in events:
            student_scores = sorted(
                [sc for sc in all_scores if sc.student_id == s.id and sc.event_id == e.id],
                key=lambda x: x.test_date
            )
            if len(student_scores) >= 2:
                prev_score = student_scores[-2].earned_score
                curr_score = student_scores[-1].earned_score
                if prev_score - curr_score >= 2:
                    warning_students.append({
                        "student_id": s.id,
                        "student_name": s.name,
                        "student_no": s.student_id,
                        "event_name": e.name,
                        "prev_score": prev_score,
                        "curr_score": curr_score
                    })

    return {
        "class_id": class_id,
        "class_name": f"{cls.grade}{cls.name}",
        "total_students": n_students,
        "avg_score": round(overall_avg, 1),
        "excellent_rate": round(excellent_count / n_students * 100, 1) if n_students else 0,
        "pass_rate": round(pass_count / n_students * 100, 1) if n_students else 0,
        "event_avgs": event_avgs,
        "warning_students": warning_students
    }

@router.get("/student-stats/{student_id}")
def student_stats(
    student_id: int,
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    s = db.query(Student).get(student_id)
    if not s:
        raise HTTPException(404, "学生不存在")

    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    scores_q = db.query(Score).filter(Score.student_id == student_id)
    if event_id_list:
        scores_q = scores_q.filter(Score.event_id.in_(event_id_list))

    all_scores = scores_q.order_by(Score.test_date.desc()).all()

    scores_by_event = defaultdict(list)
    for sc in all_scores:
        event = db.query(SportEvent).get(sc.event_id)
        scores_by_event[event.name].append({
            "id": sc.id,
            "raw_value": sc.raw_value,
            "earned_score": sc.earned_score,
            "test_date": sc.test_date.isoformat()
        })

    latest_per_event = {}
    for sc in all_scores:
        if sc.event_id not in latest_per_event:
            latest_per_event[sc.event_id] = sc

    recs = sorted(latest_per_event.items(), key=lambda x: x[1].earned_score, reverse=True)[:4]
    recommended = []
    medals = ["🥇", "🥈", "🥉", "④"]
    for i, (eid, sc) in enumerate(recs):
        event = db.query(SportEvent).get(eid)
        recommended.append({
            "rank": i + 1,
            "medal": medals[i],
            "event_name": event.name,
            "score": sc.earned_score
        })

    return {
        "student": {
            "id": s.id,
            "student_id": s.student_id,
            "name": s.name,
            "gender": s.gender.value,
            "class_name": s.class_.name if s.class_ else "",
            "class_grade": s.class_.grade if s.class_ else "",
        },
        "scores_by_event": dict(scores_by_event),
        "recommended_events": recommended
    }

@router.get("/export/class")
def export_class_scores(
    class_id: int = Query(...),
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    cls = db.query(Class).get(class_id)
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.student_id).all()
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    events_q = db.query(SportEvent)
    if event_id_list:
        events_q = events_q.filter(SportEvent.id.in_(event_id_list))
    events = events_q.order_by(SportEvent.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cls.grade}{cls.name}成绩"
    headers = ["学号", "姓名", "性别"] + [e.name for e in events] + ["总分"]
    ws.append(headers)

    for s in students:
        row = [s.student_id, s.name, "男" if s.gender.value == "M" else "女"]
        total = 0
        for e in events:
            latest = (
                db.query(Score)
                .filter(Score.student_id == s.id, Score.event_id == e.id)
                .order_by(Score.test_date.desc())
                .first()
            )
            if latest:
                row.append(latest.earned_score)
                total += latest.earned_score
            else:
                row.append("-")
        row.append(total)
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=class_{class_id}_scores.xlsx"}
    )

@router.get("/export/student/{student_id}")
def export_student_scores(
    student_id: int,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    s = db.query(Student).get(student_id)
    if not s:
        raise HTTPException(404)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "成绩汇总"
    ws1.append(["项目", "成绩", "得分", "测试日期"])
    events = db.query(SportEvent).order_by(SportEvent.sort_order).all()
    total = 0
    for e in events:
        latest = (
            db.query(Score)
            .filter(Score.student_id == student_id, Score.event_id == e.id)
            .order_by(Score.test_date.desc())
            .first()
        )
        if latest:
            ws1.append([e.name, latest.raw_value, latest.earned_score, latest.test_date.isoformat()])
            total += latest.earned_score
        else:
            ws1.append([e.name, "-", "-", "-"])
    ws1.append(["总分", "", total, ""])

    ws2 = wb.create_sheet("历史记录")
    ws2.append(["项目", "成绩", "得分", "测试日期"])
    scores = db.query(Score).filter(Score.student_id == student_id).order_by(Score.test_date.desc()).all()
    for sc in scores:
        event = db.query(SportEvent).get(sc.event_id)
        ws2.append([event.name, sc.raw_value, sc.earned_score, sc.test_date.isoformat()])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=student_{student_id}_scores.xlsx"}
    )

@router.get("/school-stats")
def school_stats(
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """School-wide statistics."""
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    events = db.query(SportEvent).order_by(SportEvent.sort_order).all()
    if event_id_list:
        events = [e for e in events if e.id in event_id_list]

    all_students = db.query(Student).all()
    all_scores = db.query(Score).order_by(Score.test_date.desc()).all()

    # Latest per student per event
    latest = {}
    for sc in all_scores:
        key = (sc.student_id, sc.event_id)
        if key not in latest:
            latest[key] = sc

    event_scores = defaultdict(list)
    student_totals = defaultdict(list)
    for (sid, eid), sc in latest.items():
        if eid in [e.id for e in events]:
            event_scores[eid].append(sc.earned_score)
            student_totals[sid].append(sc.earned_score)

    event_avgs = []
    for e in events:
        scores_list = event_scores.get(e.id, [])
        avg = sum(scores_list) / len(scores_list) if scores_list else 0
        event_avgs.append({"event_id": e.id, "event_name": e.name, "avg_score": round(avg, 1), "count": len(scores_list)})

    total_scores = [sum(v) for v in student_totals.values() if v]
    n_students = len(student_totals)
    max_per_student = len(events)
    overall_avg = sum(total_scores) / len(total_scores) if total_scores else 0
    excellent_count = sum(1 for t in total_scores if max_per_student > 0 and t / max_per_student >= 9)
    pass_count = sum(1 for t in total_scores if max_per_student > 0 and t / max_per_student >= 6)

    classes = db.query(Class).order_by(Class.grade, Class.name).all()
    class_summaries = []
    for cls in classes:
        cls_students = [s for s in all_students if s.class_id == cls.id]
        cls_total = 0
        cls_count = 0
        for s in cls_students:
            if s.id in student_totals:
                cls_total += sum(student_totals[s.id])
                cls_count += 1
        cls_avg = cls_total / (cls_count * max_per_student) * 10 if cls_count > 0 and max_per_student > 0 else 0
        class_summaries.append({
            "class_id": cls.id, "class_name": f"{cls.grade}{cls.name}",
            "students": len(cls_students), "avg_score": round(cls_avg, 1)
        })

    warning_students = []
    for s in all_students:
        for e in events:
            student_scores = sorted(
                [sc for sc in all_scores if sc.student_id == s.id and sc.event_id == e.id],
                key=lambda x: x.test_date
            )
            if len(student_scores) >= 2:
                prev = student_scores[-2].earned_score
                curr = student_scores[-1].earned_score
                if prev - curr >= 2:
                    warning_students.append({
                        "student_no": s.student_id, "student_name": s.name,
                        "event_name": e.name, "prev_score": prev, "curr_score": curr
                    })

    return {
        "total_students": n_students,
        "total_classes": len(class_summaries),
        "avg_score": round(overall_avg, 1),
        "excellent_rate": round(excellent_count / n_students * 100, 1) if n_students else 0,
        "pass_rate": round(pass_count / n_students * 100, 1) if n_students else 0,
        "event_avgs": event_avgs,
        "class_summaries": class_summaries,
        "warning_students": warning_students
    }

@router.post("/export/preview")
def export_preview(
    scope: str = Query(...),        # "school" | "class" | "student"
    class_id: Optional[int] = Query(None),
    student_id: Optional[int] = Query(None),
    event_ids: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    mode: str = Query("all"),       # "all" | "best" | "latest"
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Preview export data before downloading."""
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None

    # Build query
    q = db.query(Score)
    if scope == "class" and class_id:
        students_in_class = db.query(Student).filter(Student.class_id == class_id).all()
        q = q.filter(Score.student_id.in_([s.id for s in students_in_class]))
    elif scope == "student" and student_id:
        q = q.filter(Score.student_id == student_id)
    if event_id_list:
        q = q.filter(Score.event_id.in_(event_id_list))
    if date_from:
        q = q.filter(Score.test_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(Score.test_date <= date.fromisoformat(date_to))

    all_scores = q.order_by(Score.test_date.desc(), Score.student_id).all()
    events = db.query(SportEvent).all()
    event_map = {e.id: e for e in events}

    if mode == "best":
        best = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in best or sc.earned_score > best[key].earned_score:
                best[key] = sc
        all_scores = sorted(best.values(), key=lambda x: (x.student_id, x.event_id))
    elif mode == "latest":
        latest = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in latest:
                latest[key] = sc
        all_scores = sorted(latest.values(), key=lambda x: (x.student_id, x.event_id))

    # Preload students and classes
    student_ids = {sc.student_id for sc in all_scores} if all_scores else set()
    if student_ids:
        students_list = db.query(Student).filter(Student.id.in_(student_ids)).all()
        class_ids = {s.class_id for s in students_list}
        classes_map = {c.id: c for c in db.query(Class).filter(Class.id.in_(class_ids)).all()} if class_ids else {}
        students_map = {s.id: s for s in students_list}
    else:
        students_map = {}
        classes_map = {}

    # Build preview rows
    rows = []
    for sc in all_scores:
        student = students_map.get(sc.student_id)
        cls = classes_map.get(student.class_id) if student else None
        rows.append({
            "student_id": student.student_id if student else "",
            "student_name": student.name if student else "",
            "gender": student.gender.value if student else "",
            "class": f"{cls.grade}{cls.name}" if cls else "",
            "event_name": event_map[sc.event_id].name if sc.event_id in event_map else "",
            "raw_value": sc.raw_value,
            "earned_score": sc.earned_score,
            "test_date": sc.test_date.isoformat()
        })

    return {"rows": rows, "total": len(rows)}

@router.get("/export/download")
def export_download(
    scope: str = Query(...),
    class_id: Optional[int] = Query(None),
    student_id: Optional[int] = Query(None),
    event_ids: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    mode: str = Query("all"),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin_flexible)
):
    """Download export file (xlsx or txt)."""
    try:
        return _do_export_download(scope, class_id, student_id, event_ids, date_from, date_to, mode, format, db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")

def _do_export_download(scope, class_id, student_id, event_ids, date_from, date_to, mode, format, db):
    # Reuse preview logic
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    q = db.query(Score)
    if scope == "class" and class_id:
        students_in_class = db.query(Student).filter(Student.class_id == class_id).all()
        q = q.filter(Score.student_id.in_([s.id for s in students_in_class]))
    elif scope == "student" and student_id:
        q = q.filter(Score.student_id == student_id)
    if event_id_list:
        q = q.filter(Score.event_id.in_(event_id_list))
    if date_from:
        q = q.filter(Score.test_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(Score.test_date <= date.fromisoformat(date_to))

    all_scores = q.order_by(Score.test_date.desc(), Score.student_id).all()
    events = db.query(SportEvent).all()
    event_map = {e.id: e for e in events}

    if mode == "best":
        best = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in best or sc.earned_score > best[key].earned_score:
                best[key] = sc
        all_scores = sorted(best.values(), key=lambda x: (x.student_id, x.event_id))
    elif mode == "latest":
        latest = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in latest:
                latest[key] = sc
        all_scores = sorted(latest.values(), key=lambda x: (x.student_id, x.event_id))

    # Preload students and classes
    student_ids = {sc.student_id for sc in all_scores} if all_scores else set()
    if student_ids:
        students_list = db.query(Student).filter(Student.id.in_(student_ids)).all()
        class_ids = {s.class_id for s in students_list}
        classes_map = {c.id: c for c in db.query(Class).filter(Class.id.in_(class_ids)).all()} if class_ids else {}
        students_map = {s.id: s for s in students_list}
    else:
        students_map = {}
        classes_map = {}

    scope_label = {"school": "全校", "class": "班级", "student": "个人"}.get(scope, "")
    mode_label = {"all": "全部", "best": "最优", "latest": "最近"}.get(mode, "")

    if format == "txt":
        lines = []
        for sc in all_scores:
            student = students_map.get(sc.student_id)
            cls = classes_map.get(student.class_id) if student else None
            name = student.name if student else ""
            evt = event_map[sc.event_id].name if sc.event_id in event_map else ""
            lines.append(f"{name}\t{evt}\t{sc.raw_value}\t{sc.earned_score}分\t{sc.test_date}")
        content = "\n".join(lines)
        return StreamingResponse(
            iter([content]), media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=export_{scope}_{mode}.txt"}
        )

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{scope_label}成绩{mode_label}"
    ws.append(["学号", "姓名", "性别", "班级", "项目", "成绩", "得分", "测试日期"])
    for sc in all_scores:
        student = students_map.get(sc.student_id)
        cls = classes_map.get(student.class_id) if student else None
        ws.append([
            student.student_id if student else "", student.name if student else "",
            student.gender.value if student else "", f"{cls.grade}{cls.name}" if cls else "",
            event_map[sc.event_id].name if sc.event_id in event_map else "",
            sc.raw_value, sc.earned_score, sc.test_date.isoformat()
        ])
    buffer = BytesIO()
    wb.save(buffer); buffer.seek(0)
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=export_{scope}_{mode}.xlsx"}
    )


@router.get("/student-list/{class_id}")
def get_class_students(
    class_id: int,
    event_id: int = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    q = db.query(Student).filter(Student.class_id == class_id)
    if event_id:
        event = db.query(SportEvent).get(event_id)
        if event and event.gender.value != "both":
            q = q.filter(Student.gender == event.gender)
    students = q.order_by(Student.student_id).all()
    return [{"id": s.id, "student_id": s.student_id, "name": s.name, "gender": s.gender.value} for s in students]

@router.get("/backup-all")
def backup_all_data(db: Session = Depends(get_db), current: Admin = Depends(get_current_admin_flexible)):
    """Backup all data as Excel (7 sheets)."""
    wb = openpyxl.Workbook()

    # Sheet 1: Classes
    ws1 = wb.active
    ws1.title = "班级信息"
    ws1.append(["ID", "年级", "班级名"])
    classes = db.query(Class).order_by(Class.grade, Class.name).all()
    for c in classes:
        ws1.append([c.id, c.grade, c.name])

    # Sheet 2: Sport Events
    ws2 = wb.create_sheet("体育项目")
    ws2.append(["ID", "名称", "性别", "越大越好", "单位", "输入格式", "排序"])
    events = db.query(SportEvent).order_by(SportEvent.sort_order).all()
    for e in events:
        ws2.append([e.id, e.name, e.gender.value if e.gender else "both", e.higher_better, e.unit, e.input_format.value if e.input_format else "decimal_seconds", e.sort_order])

    # Sheet 3: Scoring Standards
    ws3 = wb.create_sheet("评分标准")
    ws3.append(["ID", "项目ID", "性别", "分数", "标准值"])
    standards = db.query(ScoringStandard).order_by(ScoringStandard.event_id, ScoringStandard.score.desc()).all()
    for std in standards:
        ws3.append([std.id, std.event_id, std.gender.value if std.gender else "both", std.score, std.standard_value])

    # Sheet 4: Students
    ws4 = wb.create_sheet("学生信息")
    ws4.append(["ID", "学号", "姓名", "性别", "班级ID", "密码哈希"])
    students = db.query(Student).order_by(Student.student_id).all()
    for s in students:
        ws4.append([s.id, s.student_id, s.name, s.gender.value if s.gender else "M", s.class_id, s.password_hash])

    # Sheet 5: Scores
    ws5 = wb.create_sheet("成绩记录")
    ws5.append(["ID", "学生ID", "项目ID", "成绩", "得分", "测试日期", "录入人ID"])
    scores = db.query(Score).order_by(Score.test_date.desc()).all()
    for sc in scores:
        ws5.append([sc.id, sc.student_id, sc.event_id, sc.raw_value, sc.earned_score, sc.test_date.isoformat(), sc.recorder_id or ""])

    # Sheet 6: Admins
    ws6 = wb.create_sheet("管理员")
    ws6.append(["ID", "用户名", "密码哈希", "超管", "显示名"])
    admins = db.query(Admin).order_by(Admin.id).all()
    for a in admins:
        ws6.append([a.id, a.username, a.password_hash, a.is_super, a.display_name])

    # Sheet 7: System Config
    ws7 = wb.create_sheet("系统设置")
    ws7.append(["键", "值"])
    configs = db.query(SystemConfig).all()
    for cfg in configs:
        ws7.append([cfg.key, cfg.value])

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sports_backup.xlsx"})

@router.post("/restore-all")
def restore_all_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Restore all data from backup Excel file."""
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(400, "无法读取文件，请确认是备份Excel文件")

    def get_sheet(name):
        if name in wb.sheetnames:
            return list(wb[name].iter_rows(min_row=2, values_only=True))
        return []

    # Read all sheets
    classes_data = [{"id": int(r[0]), "grade": str(r[1]), "name": str(r[2])} for r in get_sheet("班级信息") if r[0] is not None]
    events_data = []
    for r in get_sheet("体育项目"):
        if r[0] is not None:
            events_data.append({
                "id": int(r[0]), "name": str(r[1]), "gender": Gender(str(r[2])) if r[2] else Gender.both,
                "higher_better": r[3] in (True, "True", "true", 1, "1"),
                "unit": str(r[4]), "input_format": InputFormat(str(r[5])) if r[5] else InputFormat.decimal_seconds,
                "sort_order": int(r[6] or 0)
            })
    standards_data = []
    for r in get_sheet("评分标准"):
        if r[0] is not None:
            standards_data.append({
                "id": int(r[0]), "event_id": int(r[1]),
                "gender": Gender(str(r[2])) if r[2] else Gender.both,
                "score": int(r[3]), "standard_value": str(r[4])
            })
    students_data = []
    for r in get_sheet("学生信息"):
        if r[0] is not None:
            students_data.append({
                "id": int(r[0]), "student_id": str(r[1]), "name": str(r[2]),
                "gender": Gender(str(r[3])) if r[3] else Gender.M, "class_id": int(r[4]),
                "password_hash": str(r[5]) if len(r) > 5 and r[5] else ""
            })
    scores_data = []
    for r in get_sheet("成绩记录"):
        if r[0] is not None:
            scores_data.append({
                "id": int(r[0]), "student_id": int(r[1]), "event_id": int(r[2]),
                "raw_value": str(r[3]), "earned_score": int(r[4]),
                "test_date": date.fromisoformat(str(r[5])) if r[5] else date.today(),
                "recorder_id": int(r[6]) if r[6] and str(r[6]).strip() else None
            })
    admins_data = [{"id": int(r[0]), "username": str(r[1]), "password_hash": str(r[2]), "is_super": r[3] in (True, "True", "true", 1, "1"), "display_name": str(r[4])} for r in get_sheet("管理员") if r[0] is not None]
    configs_data = [{"key": str(r[0]), "value": str(r[1])} for r in get_sheet("系统设置") if r[0] is not None]

    try:
        # Delete in reverse dependency order
        db.query(Score).delete()
        db.query(ScoringStandard).delete()
        db.query(Student).delete()
        db.query(SportEvent).delete()
        db.query(Class).delete()
        db.query(Admin).delete()
        db.query(SystemConfig).delete()
        db.flush()

        # Insert in dependency order
        for d in classes_data:
            db.add(Class(**d))
        for d in events_data:
            db.add(SportEvent(**d))
        for d in standards_data:
            db.add(ScoringStandard(**d))
        for d in students_data:
            db.add(Student(**d))
        for d in scores_data:
            db.add(Score(**d))
        for d in admins_data:
            db.add(Admin(**d))
        for d in configs_data:
            db.add(SystemConfig(**d))
        db.flush()

        # Reset auto-increment sequences
        for tbl in ['classes', 'sport_events', 'scoring_standards', 'students', 'scores', 'admins']:
            db.execute(text(f"SELECT setval('{tbl}_id_seq', COALESCE((SELECT MAX(id) FROM {tbl}), 1))"))

        db.commit()
        return {"ok": True, "classes": len(classes_data), "events": len(events_data), "standards": len(standards_data),
                "students": len(students_data), "scores": len(scores_data), "admins": len(admins_data), "configs": len(configs_data)}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"恢复失败: {e}")

@router.post("/clear-all")
def clear_all_scores(
    data: ClearAllRequest,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_super_admin)
):
    """Delete all score records (super admin only, password required)."""
    if not verify_password(data.password, current.password_hash):
        raise HTTPException(403, "密码错误")
    count = db.query(Score).count()
    db.query(Score).delete()
    db.commit()
    return {"ok": True, "deleted": count}
